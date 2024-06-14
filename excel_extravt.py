from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# 加载Excel工作簿

wb = load_workbook(file_path)
ws = wb.active  # 假设表格数据在第一个工作表

# 调整代码以正确处理合并单元格中的单元格对象
def has_full_border(cell):
    return (cell.border.top.style is not None and
            cell.border.bottom.style is not None and
            cell.border.left.style is not None and
            cell.border.right.style is not None)

table_rows = []

for row in ws.iter_rows():
    full_border_cells_count = 0
    for cell in row:
        # 判断单元格是否位于合并单元格中
        merged = any(cell.coordinate in mr for mr in ws.merged_cells)
        if merged:
            # 从合并单元格中提取代表单元格（左上角）
            for mr in ws.merged_cells:
                if cell.coordinate in mr:
                    # 获取合并单元格的左上角单元格
                    start_cell = ws[mr.min_row][mr.min_col - 1]
                    if has_full_border(start_cell):
                        full_border_cells_count += 1
                    break
        else:
            if has_full_border(cell):
                full_border_cells_count += 1

    # 如果一行中有两个以上的合并单元格具有完整外框线
    if full_border_cells_count >= 2:
        table_rows.append([cell.value for cell in row])

# 保存数据到新Excel文件
output_path_v4 = '/mnt/data/extracted_table_with_borders_v4.xlsx'
wb_new = Workbook()
ws_new = wb_new.active
for row_data in table_rows:
    ws_new.append(row_data)
wb_new.save(output_path_v4)

output_path_v4, len(table_rows)  # 返回新文件的路径和提取的行数
